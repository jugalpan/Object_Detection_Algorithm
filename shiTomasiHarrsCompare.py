#! /usr/bin/python

#importing needed libraries
import cv2
import math
import numpy
from openpyxl import Workbook
from openpyxl.styles import Font

#defining maximum and minimum thresholds for number of corners
maxCorners = 80
minCorners = 8

#setting up Excels
harBook = Workbook()
shiBook = Workbook()

#for every image [20080, 20120] with step size n
for imageNum in range(20080, 20120, 2): # x20

	print("image: " + str(imageNum))

	#setting up Excel sheets
	harSheet = harBook.create_sheet("img" + str(imageNum))
	shiSheet = shiBook.create_sheet("img" + str(imageNum))

	#opening the image as grayscale
	gray = cv2.cvtColor(
		cv2.imread("../images/" + str(imageNum) + ".jpg"), 
		cv2.COLOR_BGR2GRAY
		)

	#for every resolution [480, 1560] with step size 120p
	#
	#testing done with only 480p; explained in report
	for res in range(480, 1680, 120): # x10

		print("\tres: " + str(res))

		#setting up Excel
		tempVert = (res / 30) - 15 #(4 * ((res / 120) - 4)) + 1

		#setting up Excel for Harris
		harSheet["A" + str(tempVert)].font = Font(bold=True)
		harSheet["A" + str(tempVert+1)].font = Font(bold=True)
		harSheet["A" + str(tempVert+2)].font = Font(bold=True)
		harSheet["A" + str(tempVert)] = "RESOLUTION"
		harSheet["A" + str(tempVert+1)] = "CONTRAST"
		harSheet["A" + str(tempVert+2)] = "AVG # KEYPTS"
		harSheet["B" + str(tempVert)] = res

		#setting up Excel for Shi-Tomasi
		shiSheet["A" + str(tempVert)].font = Font(bold=True)
		shiSheet["A" + str(tempVert+1)].font = Font(bold=True)
		shiSheet["A" + str(tempVert+2)].font = Font(bold=True)
		shiSheet["A" + str(tempVert)] = "RESOLUTION"
		shiSheet["A" + str(tempVert+1)] = "CONTRAST"
		shiSheet["A" + str(tempVert+2)] = "AVG # KEYPTS"
		shiSheet["B" + str(tempVert)] = res

		#spatially rescale the image at 16:9/w:h ratio
		scaled = cv2.resize(gray, (int(math.ceil(res * 16 / 9.0)), res))

		#for every contrast [0.01, 0.31] at step size 0.02
		for contraThresh in range(1, 33, 2): # x16
			#Harris and Shi-Tomasi can run within this

			#changing the contrast to something usable
			tempContraThresh = contraThresh / 100.0
			print("\t\tcontrast: " + str(tempContraThresh))

			#setting up Excel sheets
			tempCol = chr((contraThresh/2) + (contraThresh%2) + 65)

			#variables for calculating the averages
			harTempSum = 0
			harCounter = 0
			harSheet[str(tempCol) + str(tempVert+1)] = tempContraThresh

			shiTempSum = 0
			shiCounter = 0
			shiSheet[str(tempCol) + str(tempVert+1)] = tempContraThresh

			#for every Sobel [1, 9] at step size 2
			for sobel in range(1, 11, 2): # x5

				#for every Harris corner threshold [0.5, 1.0] at step 
					# size 0.05
				for cornerThresh in range(64, 100, 2): # x18
					#Harris is run here

					#changing the threshold to something usable
					tempCornerThresh = cornerThresh / 100.0

					#calculating the corner-ness of each pixel
					harCorners = cv2.dilate(
						cv2.cornerHarris(
							k=tempContraThresh,
							ksize=sobel,

							#blockSize is kept constant
							blockSize=3,
							src=numpy.float32(scaled)
							),
						None
						)

					#thresholding the Harris measure
					harCompar = tempCornerThresh * harCorners.max()
					harCorners = harCorners[harCorners > harCompar]

					#thresholding by number of corners
					if (minCorners < len(harCorners) < maxCorners):
						harTempSum += len(harCorners)
						harCounter += 1
					#if
				#for cornerThresh in range (x18)
			#for sobel in range (x5)

			#for every minimum corner distance [5, 25] st step size 5
			for minDist in range(5, 30, 5): # x5

				#for every quality [0.01, 0.40] at step size 4
				for quality in range(1, 11, 1): # x10
					#Shi-Tomasi is run here

					#changing the quality to something usable
					tempQual = quality / 100.0

					#calculating the corners
					shiCorners = numpy.int_(
							cv2.goodFeaturesToTrack(
							k=tempContraThresh,
							qualityLevel=tempQual,
							minDistance=minDist,

							#maxCorners is kept constant
							maxCorners=maxCorners,
							#blockSize is kept constant
							blockSize=3,
							image=scaled 
							)
						)

					#thresholding by number of corners
					if (minCorners < len(shiCorners) < maxCorners):
						shiTempSum += len(shiCorners)
						shiCounter += 1
					#if

				#for quality in range (x5)
			#for minDist in range (x10) 

			#making sure there is no divide by zero error
			if (harCounter == 0):
				harCounter += 1
			#if
			if (shiCounter == 0):
				shiCounter += 1
			#if

			#adding the average to the Excel
			harSheet[str(tempCol) + str(tempVert+2)] = (harTempSum / 
				harCounter)
			shiSheet[str(tempCol) + str(tempVert+2)] = (shiTempSum / 
				shiCounter)
		#for contraThresh in range (x15)

		#periodically saving the Excel sheets
		harBook.save("Harris.xlsx")
		shiBook.save("Shi-Tomasi.xlsx")

		break
	#for res in range (x10)
#for imageNum in range (x10)

print("Program complete.")