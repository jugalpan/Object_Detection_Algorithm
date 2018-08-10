#! /usr/bin/python

import sys
import cv2
import math
from openpyxl import Workbook
from openpyxl.styles import Font

maxKeypoints = 3000
minKeypoints = 30
maxCorners = 800
minCorners = 8

siftSigma = 1.6 #according to Lowe's paper

siftBook = Workbook()
orbBook = Workbook()
surfBook = Workbook()

#for every image [20080, 20120] with step size n
for imageNum in range(20080, 20090, 2): # x40

	print("image: " + str(imageNum))

	#setting up Excel sheets
	siftSheet = siftBook.create_sheet("img" + str(imageNum))
	orbSheet = orbBook.create_sheet("img" + str(imageNum))
	surfSheet = surfBook.create_sheet("img" + str(imageNum))

	gray = cv2.cvtColor(
		cv2.imread("../images/" + str(imageNum) + ".jpg"), 
		cv2.COLOR_BGR2GRAY
		)

	#for every resolution [480, 1560] with step size 120p
	for res in range(480, 1680, 120): # x10

		print("\tres: " + str(res))

		tempVert = (res / 30) - 15 #(4 * ((res / 120) - 4)) + 1

		#setting up Excel sheet for SIFT
		siftSheet["A" + str(tempVert)].font = Font(bold=True)
		siftSheet["A" + str(tempVert+1)].font = Font(bold=True)
		siftSheet["A" + str(tempVert+2)].font = Font(bold=True)
		siftSheet["A" + str(tempVert)] = "RESOLUTION"
		siftSheet["A" + str(tempVert+1)] = "INTV / OCTV"
		siftSheet["A" + str(tempVert+2)] = "AVG # KEYPTS"
		siftSheet["B" + str(tempVert)] = res

		#setting up Excel sheet for ORB		
		orbSheet["A" + str(tempVert)].font = Font(bold=True)
		orbSheet["A" + str(tempVert+1)].font = Font(bold=True)
		orbSheet["A" + str(tempVert+2)].font = Font(bold=True)
		orbSheet["A" + str(tempVert)] = "RESOLUTION"
		orbSheet["A" + str(tempVert+1)] = "INTV / OCTV"
		orbSheet["A" + str(tempVert+2)] = "AVG # KEYPTS"
		orbSheet["B" + str(tempVert)] = res

		#setting up Excel sheet for SURF
		surfSheet["A" + str(tempVert)].font = Font(bold=True)
		surfSheet["A" + str(tempVert+1)].font = Font(bold=True)
		surfSheet["A" + str(tempVert+2)].font = Font(bold=True)
		surfSheet["A" + str(tempVert)] = "RESOLUTION"
		surfSheet["A" + str(tempVert+1)] = "INTV / OCTV"
		surfSheet["A" + str(tempVert+2)] = "AVG # KEYPTS"
		surfSheet["B" + str(tempVert)] = res

		#spatially rescale the image at 16:9 w:h ratio
		scaled = cv2.resize(gray, (int(math.ceil(res * 16 / 9.0)), res))

		#doesn't make sense look at images smaller than 
			# <math.floor(math.log(res, 2)) - 3> x 
			# <math.floor(math.log(res, 2)) - 3>
		#used to keep close to SIFT, which also bases number of octaves
			#on the size of the image
		surfNumOctaves = int(math.floor(math.log(res, 2)) - 3)

		#for every number of intervals [1, 13] with step size 1
		for numIntervals in range(1, 14): # x13
			#SIFT, SURF, and ORB can run within this

			print("\t\tnumIntervals: " + str(numIntervals))

			#setting up Excel sheets
			tempCol = chr(numIntervals+65)

			siftTempSum = 0
			siftCounter = 0			
			siftSheet[str(tempCol) + str(tempVert+1)] = numIntervals

			orbTempSum = 0
			orbCounter = 0			
			orbSheet[str(tempCol) + str(tempVert+1)] = numIntervals

			surfTempSum = 0
			surfCounter = 0			
			surfSheet[str(tempCol) + str(tempVert+1)] = numIntervals

			#for every edge threhsold [0, 24] with step size 2
			for edgeThresh in range(0, 26, 2): # x13
				#SIFT and ORB can run within this

				#for every contrast threshold [0.02, 0.10] step size 0.2
				for contraThresh in range(2, 12, 2): # x6
					#SIFT is run here

					sift = cv2.xfeatures2d.SIFT_create(
						nfeatures=maxKeypoints,
						nOctaveLayers=numIntervals,
						edgeThreshold=edgeThresh,
						contrastThreshold=contraThresh/100.0,

						#sigma is kept constant
						sigma=siftSigma
						)
					siftKeys = sift.detect(scaled, None)
					numSiftKeys = len(siftKeys)
					if (numSiftKeys >= minKeypoints):
						siftTempSum += numSiftKeys
						siftCounter += 1
					#if
				#for contraThresh in range (x6)

				#for every FAST threshold [1, 41] with step size 10
				orbEdgeThresh = edgeThresh * 10
				for fastThresh in range(1, 51, 10): # x5
					#ORB is run here

					orb = cv2.ORB_create(
						nfeatures=maxKeypoints,
						nlevels=numIntervals,
						edgeThreshold=orbEdgeThresh,
						fastThreshold=fastThresh
						)
					orbKeys = orb.detect(scaled, None)
					numOrbKeys = len(orbKeys)
					if (numOrbKeys >= minKeypoints):
						orbTempSum += numOrbKeys
						orbCounter += 1
					#if
				#for fastThresh in range (x5)
			#for edgeThresh in range (x13)

			#for every Hessian threshold [295, 505] with step size 3
			for hessThresh in range(295, 508, 3): # x70
				#SURF is run here

				surf = cv2.xfeatures2d.SURF_create(
					nOctaveLayers=numIntervals,
					hessianThreshold=hessThresh,

					#nOctaves is kept constant
					nOctaves=surfNumOctaves
					)
				surfKeys = surf.detect(scaled, None)
				numSurfKeys = len(surfKeys)
				if (minKeypoints <= numSurfKeys <= maxKeypoints):
					surfTempSum += numSurfKeys
					surfCounter += 1
				#if
			#for hessThresh in range (x70)

			# print("\t\t\tavg numSiftKeys:\t" + str(siftTempSum / 
			# 	siftCounter))
			# print("\t\t\tavg numOrbKeys:\t\t" + str(orbTempSum / 
			# 	orbCounter))
			if (siftCounter == 0):
				siftCounter += 1
			#if
			if (orbCounter == 0):
				orbCounter += 1
			#if
			if (surfCounter == 0):
				surfCounter += 1
			#if

			siftSheet[str(tempCol) + str(tempVert+2)] = (siftTempSum / 
				siftCounter)
			orbSheet[str(tempCol) + str(tempVert+2)] = (orbTempSum / 
				orbCounter)
			surfSheet[str(tempCol) + str(tempVert+2)] = (surfTempSum / 
				surfCounter)
		#for numIntervals in range (x13)
		siftBook.save("SIFT.xlsx")
		orbBook.save("ORB.xlsx")
		surfBook.save("SURF.xlsx")

		break
	#for res in range (x10)
	# break
#for imageNum in range (x10)

print("Program complete.")